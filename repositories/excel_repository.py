from pathlib import Path
from openpyxl import Workbook, load_workbook
from models.CheckInRecord import CheckInRecord


class ExcelRepository:
    HEADERS = [
        "Name",
        "Phone",
        "Reason",
        "Check In Time",
        "Photo Path",
    ]

    # initializes the ExcelRepository with the given file path
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self._initialize_workbook()

    # initializes the Excel workbook if it doesn't exist
    def _initialize_workbook(self):
        if self.file_path.exists():
            return

        self.file_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Check Ins"

        sheet.append(self.HEADERS)

        workbook.save(self.file_path)

    # saves a CheckInRecord object to the Excel file
    def save(self, record: CheckInRecord):
        workbook = load_workbook(self.file_path)
        sheet = workbook["Check Ins"]

        sheet.append([
            record.name,
            record.phone,
            record.reason,
            record.check_in_time,
            record.photo_path,
        ])

        workbook.save(self.file_path)
        workbook.close()

    # gets all check-in records from the Excel file and returns them as a list of CheckInRecord objects
    def get_all(self) -> list[CheckInRecord]:
        workbook = load_workbook(self.file_path)
        sheet = workbook["Check Ins"]

        records = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            record = CheckInRecord(
                name=row[0],
                phone=row[1],
                reason=row[2],
                check_in_time=row[3],
                photo_path=row[4],
            )

            records.append(record)

        workbook.close()

        return records
